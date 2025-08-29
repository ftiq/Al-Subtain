# -*- coding: utf-8 -*-

import base64
import io
import json
from datetime import datetime
import logging
import re
from odoo import models, fields, api, _
from odoo.exceptions import UserError
from werkzeug.urls import url_quote
from odoo import Command

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.drawing.image import Image as ExcelImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

_logger = logging.getLogger(__name__)


class LabSampleExcelExportWizard(models.TransientModel):
    """معالج تصدير نتائج عينة المختبر إلى Excel"""
    _name = 'lab.sample.excel.export.wizard'
    _description = 'تصدير نتائج عينة المختبر إلى Excel'

    sample_id = fields.Many2one(
        'lab.sample', 
        string='العينة', 
        required=True, 
        readonly=True
    )
    include_summary = fields.Boolean(
        string='تضمين ملخص النتائج', 
        default=True,
        help="تضمين ملخص إجمالي للنتائج في أعلى التقرير"
    )
    include_header_image = fields.Boolean(
        string='تضمين صورة الترويسة',
        default=True,
        help="تضمين صورة الترويسة من إعدادات الشركة"
    )
    include_footer_image = fields.Boolean(
        string='تضمين صورة التذييل',
        default=True,
        help="تضمين صورة التذييل من إعدادات الشركة"
    )

    group_by_template = fields.Boolean(default=True)


    sample_subtype_id = fields.Many2one(
        related='sample_id.sample_subtype_id',
        string='النوع الفرعي',
        readonly=True,
        store=False,
    )

    # تكامل Spreadsheet
    current_spreadsheet_attachment_id = fields.Many2one('ir.attachment', string='الملف المفتوح')

    def action_open_spreadsheet(self):
        self.ensure_one()
        if not self.sample_id:
            raise UserError(_("يجب تحديد عينة أولاً"))



        session = self.env['lab.spreadsheet.session'].create({
            'name': f"Spreadsheet - {self.sample_id.name}",
        })

        session.spreadsheet_binary_data = self.env['spreadsheet.mixin']._empty_spreadsheet_data_base64()
        self.current_spreadsheet_attachment_id = False

        # Inject lists into spreadsheet session
        if False:
            try:
                # Collect all lists in the template that work on lab.result.line model
                snapshot = document._get_spreadsheet_snapshot()
                lists = snapshot.get('lists') or {}
                mapping = {}
                for list_id, lst in lists.items():
                    try:
                        if lst.get('model') == 'lab.result.line':
                            mapping[str(list_id)] = {'chain': 'result_set_id.sample_id', 'type': 'many2one'}
                    except Exception:
                        continue


                if mapping:
                    add_cmd = {
                        'type': 'ADD_GLOBAL_FILTER',
                        'filter': {
                            'id': 'sample_filter_id',
                            'type': 'relation',
                            'label': 'العينة',
                            'modelName': 'lab.sample',
                            'defaultValue': [self.sample_id.id],
                            'defaultValueDisplayNames': [self.sample_id.display_name],
                        },
                        'list': mapping,
                    }
                    try:
                        document._dispatch_commands([add_cmd])
                    except Exception:
                        edit_cmd = {
                            'type': 'EDIT_GLOBAL_FILTER',
                            'filter': {'id': 'sample_filter_id'},
                            'list': mapping,
                        }
                        document._dispatch_commands([edit_cmd])


                set_filter_value_cmd = {
                    'type': 'SET_GLOBAL_FILTER_VALUE',
                    'id': 'sample_filter_id',
                    'value': [self.sample_id.id],
                    'displayNames': [self.sample_id.display_name],
                }
                document._dispatch_commands([set_filter_value_cmd])
            except Exception:
                pass
        else:

            try:
                ResultLine = self.env['lab.result.line']

                list_columns = [
                    'sample_identifier', 'criterion_name', 'unit_of_measure',
                    'min_value', 'max_value', 'value_numeric', 'result_value_computed',
                    'conformity_status', 'notes'
                ]

                result_sets = self.sample_id.result_set_ids.sorted(lambda rs: (rs.template_id.name or '', rs.id))
                row_cursor = 0
                next_list_id = 1
                for rs in result_sets:

                    title = f"مجموعة النتائج: {rs.name or ''} — قالب: {(rs.template_id.name or 'بدون قالب')}"
                    title_cmd = {
                        'type': 'UPDATE_CELL',
                        'sheetId': 'sheet1',
                        'row': row_cursor,
                        'col': 0,
                        'content': title,
                        'style': {'bold': True},
                    }
                    row_cursor += 1

                    lines_count = len(rs.result_line_ids) or 0
                    lines_num = max(lines_count + 5, 20)  
                    list_id = str(next_list_id)
                    next_list_id += 1

                    definition = {
                        'metaData': {
                            'resModel': 'lab.result.line',
                            'columns': list_columns,
                        },
                        'searchParams': {
                            'domain': [["result_set_id", "=", rs.id]],
                            'context': {},
                            'orderBy': [],
                        },
                        'limit': lines_num,
                        'name': rs.display_name or rs.name or _('Lines'),
                    }
                    insert_cmd = {
                        'type': 'INSERT_ODOO_LIST',
                        'sheetId': 'sheet1',
                        'col': 0,
                        'row': row_cursor,
                        'id': list_id,
                        'definition': definition,
                        'linesNumber': lines_num,
                        'columns': [
                            {'name': col, 'type': ResultLine._fields[col].type if col in ResultLine._fields else 'char'}
                            for col in list_columns
                        ],
                    }
                    try:

                        session._dispatch_commands([title_cmd, insert_cmd])
                    except Exception:
                        pass


                    row_cursor += lines_num + 2
            except Exception:
                pass

        return session.action_open_spreadsheet()

    def action_save_current_as_template(self):
        self.ensure_one()
        if not self.current_spreadsheet_attachment_id:
            raise UserError(_("لا يوجد ملف Spreadsheet مفتوح للحفظ كقالب"))
        self.env['lab.spreadsheet.template'].create({
            'name': f"قالب - {self.sample_id.name}",
            'attachment_id': self.current_spreadsheet_attachment_id.id,
            'sample_type_id': self.sample_id.product_id.product_tmpl_id.sample_type_id.id if self.sample_id.product_id and self.sample_id.product_id.product_tmpl_id.sample_type_id else False,
            'sample_subtype_id': self.sample_id.sample_subtype_id.id if self.sample_id.sample_subtype_id else False,
        })
        return {'type': 'ir.actions.client', 'tag': 'display_notification', 'params': {'type': 'success', 'message': _('تم حفظ القالب') }}

    def action_export_via_spreadsheet(self):
        self.ensure_one()
        if not self.current_spreadsheet_attachment_id:
            raise UserError(_("لا يوجد Spreadsheet للتصدير"))

        doc = self.env['documents.document'].sudo().search([('attachment_id', '=', self.current_spreadsheet_attachment_id.id)], limit=1)
        if not doc:
            raise UserError(_("المستند غير موجود"))

        if doc.excel_export:
            att = self.env['ir.attachment'].create({
                'name': f"Spreadsheet_{self.sample_id.name}.xlsx",
                'type': 'binary',
                'datas': doc.excel_export,
                'res_model': self._name,
                'res_id': self.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })
            return {'type': 'ir.actions.act_url', 'url': f"/web/content/{att.id}?download=1", 'target': 'self'}
        return {'type': 'ir.actions.client', 'tag': 'display_notification', 'params': {'type': 'warning', 'message': _('يرجى التصدير من داخل المحرر (ملف > تنزيل كـ XLSX)')}}

    export_file = fields.Binary(
        string='ملف Excel المُصدر', 
        readonly=True
    )
    export_filename = fields.Char(
        string='اسم الملف', 
        readonly=True
    )
    state = fields.Selection([
        ('draft', 'معاينة'),
        ('exported', 'تم التصدير')
    ], default='draft', string='الحالة')
    
    # computed fields
    sample_lab_code = fields.Char(
        string='الرمز المختبري',
        compute='_compute_sample_info'
    )
    sample_product_name = fields.Char(
        string='اسم المنتج',
        compute='_compute_sample_info'
    )
    result_sets_count = fields.Integer(
        string='عدد مجموعات النتائج',
        compute='_compute_sample_info'
    )
    result_lines_count = fields.Integer(
        string='إجمالي خطوط النتائج',
        compute='_compute_sample_info'
    )
    
    @api.depends('sample_id')
    def _compute_sample_info(self):
        """حساب معلومات العينة للعرض"""
        for wizard in self:
            if wizard.sample_id:
                wizard.sample_lab_code = wizard.sample_id.lab_code or ''
                wizard.sample_product_name = wizard.sample_id.product_id.name if wizard.sample_id.product_id else ''
                wizard.result_sets_count = len(wizard.sample_id.result_set_ids)
                result_lines_count = 0
                for rs in wizard.sample_id.result_set_ids:
                    result_lines_count += len(rs.result_line_ids)
                wizard.result_lines_count = result_lines_count
            else:
                wizard.sample_lab_code = ''
                wizard.sample_product_name = ''
                wizard.result_sets_count = 0
                wizard.result_lines_count = 0

    @api.model
    def default_get(self, fields_list):
        """تعيين العينة من السياق"""
        res = super().default_get(fields_list)
        sample_id = self.env.context.get('active_id')
        if sample_id:
            res['sample_id'] = sample_id
        return res

    def action_export_excel(self):
        """تصدير البيانات إلى Excel"""
        self.ensure_one()
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("مكتبة openpyxl غير مثبتة. يرجى تثبيتها أولاً: pip install openpyxl"))
        
        if not PIL_AVAILABLE and (self.include_header_image or self.include_footer_image):
            _logger.warning("مكتبة PIL (Pillow) غير متوفرة. قد تكون جودة الصور أقل. ينصح بتثبيتها: pip install pillow")
        
        if not self.sample_id:
            raise UserError(_("يجب تحديد عينة للتصدير"))


        if False and self.current_spreadsheet_attachment_id:

            pass




        wb = openpyxl.Workbook()
        

        wb.remove(wb.active)
        
        ws_main = wb.create_sheet(title="نتائج العينة")
        
        self._setup_worksheet_rtl(ws_main)
        
        current_row = 1
        
        # Add header image if available
        if self.include_header_image:
            current_row = self._add_header_image(ws_main, current_row)
        
        current_row = self._add_main_header(ws_main, current_row)
        
        current_row = self._add_sample_info(ws_main, current_row)
        
        if self.include_summary:
            current_row = self._add_results_summary(ws_main, current_row)
        
        current_row = self._add_detailed_results(ws_main, current_row)
        
        # إضافة صورة التذييل إذا كانت متوفرة
        if self.include_footer_image:
            current_row = self._add_footer_image(ws_main, current_row)
        

        
        self._finalize_sheet(ws_main)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        content_bytes = output.getvalue()
        content_b64 = base64.b64encode(content_bytes)

        display_filename = f"نتائج_العينة_{self.sample_id.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        safe_filename = f"sample_{self.sample_id.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.write({
            'export_file': content_b64,
            'export_filename': display_filename,
            'state': 'exported'
        })

        Attachment = self.env['ir.attachment'].sudo().create({
            'name': safe_filename,
            'type': 'binary',
            'datas': base64.b64encode(content_bytes),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        download_url = f"/web/content/{Attachment.id}?download=1&filename={url_quote(safe_filename)}"
        return {
            'type': 'ir.actions.act_url',
            'url': download_url,
            'target': 'self',
        }

    def _setup_worksheet_rtl(self, worksheet):

        worksheet.sheet_view.rightToLeft = True
        worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        try:
            worksheet.page_setup.fitToWidth = 1
            worksheet.page_setup.fitToHeight = 0
        except Exception:
            pass
        try:
            worksheet.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2)
        except Exception:
            pass

    def _add_header_image(self, worksheet, start_row):
        """إضافة صورة الترويسة من إعدادات الشركة"""
        try:
            company = self.env.company
            if not company.header_image:
                return start_row
            
            # تحويل binary data إلى BytesIO
            import base64
            image_data = base64.b64decode(company.header_image)
            image_stream = io.BytesIO(image_data)
            
            # إنشاء كائن الصورة لـ openpyxl
            img = ExcelImage(image_stream)
            
            # تحديد حجم الصورة (تقريبياً 20 سم عرض × 3 سم ارتفاع)
            # 1 سم ≈ 37.8 pixels في Excel عند 96 DPI
            # تحويل لوحدات Excel EMU (English Metric Units)
            # 1 pixel = 9525 EMU, 1 سم = 360000 EMU
            target_width_cm = 20
            target_height_cm = 3
            
            if PIL_AVAILABLE:
                # إعادة تعيين المؤشر لبداية الملف
                image_stream.seek(0)
                # فتح الصورة باستخدام PIL لضبط الحجم بدقة
                pil_img = PILImage.open(image_stream)
                # تحويل إلى RGB إذا كانت بصيغة RGBA
                if pil_img.mode in ('RGBA', 'LA'):
                    background = PILImage.new('RGB', pil_img.size, (255, 255, 255))
                    background.paste(pil_img, mask=pil_img.split()[-1] if pil_img.mode == 'RGBA' else None)
                    pil_img = background
                
                # حفظ الصورة في مجرى BytesIO جديد
                processed_stream = io.BytesIO()
                pil_img.save(processed_stream, format='PNG', quality=95, optimize=True)
                processed_stream.seek(0)
                
                # إنشاء كائن الصورة من المجرى المعالج
                img = ExcelImage(processed_stream)
            else:
                img = ExcelImage(image_stream)
            

            img.width = int(target_width_cm * 37.8)  # 20 سم = ~756 pixels
            img.height = int(target_height_cm * 37.8)  # 3 سم = ~113 pixels
            
            # إدراج الصورة في الخلية A1 مع توسيطها
            worksheet.add_image(img, f'A{start_row}')
            
            # إضافة صفوف فارغة لمساحة الصورة
            return start_row + 4
            
        except Exception as e:
            _logger.warning(f"خطأ في إضافة صورة الترويسة: {e}")
            return start_row
    
    def _add_footer_image(self, worksheet, start_row):
        """إضافة صورة التذييل من إعدادات الشركة"""
        try:
            company = self.env.company
            if not company.footer_image:
                return start_row
            
            # إضافة مساحة فارغة قبل التذييل
            start_row += 2
            
            # تحويل binary data إلى BytesIO
            import base64
            image_data = base64.b64decode(company.footer_image)
            image_stream = io.BytesIO(image_data)
            
            # إنشاء كائن الصورة لـ openpyxl
            if PIL_AVAILABLE:
                # إعادة تعيين المؤشر لبداية الملف
                image_stream.seek(0)
                pil_img = PILImage.open(image_stream)
                if pil_img.mode in ('RGBA', 'LA'):
                    background = PILImage.new('RGB', pil_img.size, (255, 255, 255))
                    background.paste(pil_img, mask=pil_img.split()[-1] if pil_img.mode == 'RGBA' else None)
                    pil_img = background
                
                processed_stream = io.BytesIO()
                pil_img.save(processed_stream, format='PNG', quality=95, optimize=True)
                processed_stream.seek(0)
                img = ExcelImage(processed_stream)
            else:
                img = ExcelImage(image_stream)
            
            # تحديد حجم الصورة (تقريبياً 20 سم عرض × 3 سم ارتفاع)
            img.width = int(20 * 37.8)  # 20 سم = ~756 pixels
            img.height = int(3 * 37.8)  # 3 سم = ~113 pixels
            
            # إدراج الصورة في الخلية
            worksheet.add_image(img, f'A{start_row}')
            
            # إضافة صفوف فارغة لمساحة الصورة
            return start_row + 4
            
        except Exception as e:
            _logger.warning(f"خطأ في إضافة صورة التذييل: {e}")
            return start_row
    
    def _add_main_header(self, worksheet, start_row):

        worksheet.merge_cells(f'A{start_row}:H{start_row}')
        cell = worksheet[f'A{start_row}']
        cell.value = f"تقرير نتائج فحص العينة - {self.sample_id.name}"
        cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        start_row += 1
        worksheet.merge_cells(f'A{start_row}:H{start_row}')
        cell = worksheet[f'A{start_row}']
        cell.value = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        cell.font = Font(name='Arial', size=12, italic=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        return start_row + 2

    def _add_sample_info(self, worksheet, start_row):
        """إضافة معلومات العينة"""

        worksheet.merge_cells(f'A{start_row}:H{start_row}')
        cell = worksheet[f'A{start_row}']
        cell.value = "معلومات العينة"
        cell.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        start_row += 1
        
        task = self.sample_id.task_id
        sample_info = [
            ('رقم العينة:', self.sample_id.name or ''),
            ('المنتج:', self.sample_id.product_id.name or ''),
            ('نوع عينة الخرسانة:', (self.sample_id.concrete_sample_type_id.name
                                   if getattr(self.sample_id, 'concrete_sample_type_id', False) and self.sample_id.concrete_sample_type_id
                                   else '')),
            ('النوع الفرعي:', (f"{self.sample_id.sample_subtype_id.name} ({self.sample_id.sample_subtype_id.code})"
                               if self.sample_id.sample_subtype_id and self.sample_id.sample_subtype_id.code
                               else (self.sample_id.sample_subtype_id.name if self.sample_id.sample_subtype_id else ''))),
            ('الرمز المختبري:', self.sample_id.lab_code or ''),
            ('تاريخ الجمع:', self.sample_id.collection_date.strftime('%Y-%m-%d') if self.sample_id.collection_date else ''),
            ('تاريخ الاستلام:', self.sample_id.received_date.strftime('%Y-%m-%d') if self.sample_id.received_date else ''),
            ('الكمية:', f"{self.sample_id.quantity or 0}"),
            ('الحالة:', dict(self.sample_id._fields['state'].selection).get(self.sample_id.state, '')),

            ('المشروع:', task.project_id.name if task and task.project_id else ''),
            ('المهمة:', task.name if task else ''),
            ('عرض السعر:', self.sample_id.sale_order_id.name if getattr(self.sample_id, 'sale_order_id', False) and self.sample_id.sale_order_id else ''),
            ('العميل:', self.sample_id.partner_id.name if getattr(self.sample_id, 'partner_id', False) and self.sample_id.partner_id else ''),
            ('رقم الكتاب:', task.book_number if task and hasattr(task, 'book_number') else ''),
            ('تاريخ الكتاب:', task.book_date.strftime('%Y-%m-%d') if task and getattr(task, 'book_date', False) else ''),
            ('تاريخ النمذجة:', task.modeling_date.strftime('%Y-%m-%d') if task and getattr(task, 'modeling_date', False) else ''),
            ('الملاحظات:', self._to_plain_text(task.book_notes) if task and getattr(task, 'book_notes', False) else ''),
        ]

        pairs_per_row = 3
        label_columns = [1, 3, 5]
        info_label_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        for i, (label, value) in enumerate(sample_info):
            row = start_row + (i // pairs_per_row)
            col = label_columns[i % pairs_per_row]
            cell_label = worksheet.cell(row=row, column=col)
            cell_value = worksheet.cell(row=row, column=col + 1)
            cell_label.value = label
            cell_value.value = value
            cell_label.font = Font(name='Arial', size=11, bold=True)
            cell_value.font = Font(name='Arial', size=11)
            cell_label.alignment = Alignment(horizontal='right', vertical='center')
            cell_value.alignment = Alignment(horizontal='center', vertical='center')
            cell_label.fill = info_label_fill

        used_rows = (len(sample_info) + pairs_per_row - 1) // pairs_per_row
        return start_row + used_rows + 1

    def _add_results_summary(self, worksheet, start_row):
        """إضافة ملخص النتائج"""
        worksheet.merge_cells(f'A{start_row}:H{start_row}')
        cell = worksheet[f'A{start_row}']
        cell.value = "ملخص النتائج"
        cell.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='9BBB59', end_color='9BBB59', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        start_row += 1
        
        task = getattr(self.sample_id, 'task_id', False)
        try:
            ref_general = float(task.reference_general_limit) if task and getattr(task, 'reference_general_limit', False) else None
        except Exception:
            ref_general = None
        try:
            ref_min = float(task.reference_min_limit) if task and getattr(task, 'reference_min_limit', False) else None
        except Exception:
            ref_min = None
        # قراءة هامش المعيار العام من الإعدادات
        try:
            general_margin = float(self.env['ir.config_parameter'].sudo().get_param('appointment_products.general_limit_margin', default='3'))
        except Exception:
            general_margin = 3.0

        try:
            min_margin = float(self.env['ir.config_parameter'].sudo().get_param('appointment_products.min_limit_margin', default='3'))
        except Exception:
            min_margin = 3.0
        is_bridges = False
        try:
            st = getattr(self.sample_id, 'concrete_sample_type_id', False)
            is_bridges = bool(st and getattr(st, 'code', '') == 'CONCRETE_BRIDGES')
        except Exception:
            is_bridges = False

        all_sets = self.sample_id.result_set_ids
        included_sets = all_sets.filtered(
            lambda rs: (not getattr(rs, 'is_concrete_sample', False)) or (getattr(rs, 'concrete_age_days', False) == '28')
        )
        total_result_sets = len(included_sets)
        completed_sets = len(included_sets.filtered(lambda x: x.state in ['approved', 'completed']))
        passed_sets = len(included_sets.filtered(lambda x: x.overall_result == 'pass'))
        failed_sets = len(included_sets.filtered(lambda x: x.overall_result == 'fail'))

        group_failed = False
        rolling_failed = False
        try:
            concrete_sets_28 = self.sample_id.result_set_ids.filtered(
                lambda rs: getattr(rs, 'is_concrete_sample', False) and getattr(rs, 'concrete_age_days', False) == '28'
            ).sorted('concrete_group_no')
            group_avgs = []
            for rs in concrete_sets_28:
                line = rs.result_line_ids.filtered(lambda l: l.criterion_id.code == 'AVG_COMP_STRENGTH_CONCRETE')
                if line:
                    grp_avg = line[0].value_numeric
                else:
                    strengths = self._get_values_for_set(rs, 'COMP_STRENGTH_CONCRETE')
                    grp_avg = (sum(strengths)/len(strengths)) if strengths else 0
                group_avgs.append(grp_avg)
                if ref_min is not None and grp_avg < (ref_min - min_margin):
                    group_failed = True
            if not is_bridges and ref_general is not None and len(group_avgs) >= 4:
                window = 4
                for i in range(0, len(group_avgs) - window + 1):
                    avg_val = sum(group_avgs[i:i+window]) / window

                    if avg_val < (ref_general + general_margin):
                        rolling_failed = True
                        break
        except Exception:
            pass

        final_failed = group_failed or rolling_failed
        final_result_text = 'فاشل' if final_failed else 'ناجح'

        summary_info = [
            ('إجمالي المجموعات:', str(total_result_sets)),
            ('المكتملة:', str(completed_sets)),
            ('الناجحة:', str(passed_sets)),
            ('الفاشلة:', str(failed_sets)),
            ('النتيجة الإجمالية:', final_result_text),
            ('نسبة التقدم:', f"{self.sample_id.progress_percentage}%" if self.sample_id.progress_percentage else '0%'),
        ]

        pairs_per_row = 3
        label_columns = [1, 3, 5]
        summary_label_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        for i, (label, value) in enumerate(summary_info):
            row = start_row + (i // pairs_per_row)
            col = label_columns[i % pairs_per_row]
            cell_label = worksheet.cell(row=row, column=col)
            cell_value = worksheet.cell(row=row, column=col + 1)
            cell_label.value = label
            cell_value.value = value
            cell_label.font = Font(name='Arial', size=11, bold=True)
            cell_value.font = Font(name='Arial', size=11)
            cell_label.alignment = Alignment(horizontal='right', vertical='center')
            cell_value.alignment = Alignment(horizontal='center', vertical='center')
            cell_label.fill = summary_label_fill

            if label == 'الناجحة:':
                cell_value.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif label == 'الفاشلة:':
                cell_value.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif label == 'النتيجة الإجمالية:':
                if final_failed:
                    cell_value.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                else:
                    cell_value.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        used_rows = (len(summary_info) + pairs_per_row - 1) // pairs_per_row
        return start_row + used_rows + 1

    def _add_detailed_results(self, worksheet, start_row):
        """إضافة النتائج التفصيلية"""
        worksheet.merge_cells(f'A{start_row}:H{start_row}')
        cell = worksheet[f'A{start_row}']
        cell.value = "النتائج التفصيلية"
        cell.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='F79646', end_color='F79646', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        start_row += 2
        

        if self._is_brick_sample():
            start_row = self._add_brick_consolidated_report(worksheet, start_row)
            return start_row

        if self._is_concrete_sample():
            start_row = self._add_concrete_compression_report(worksheet, start_row)
            return start_row

        if self.group_by_template:
            result_sets_grouped = {}
            for idx_rs, result_set in enumerate(self.sample_id.result_set_ids.sorted(lambda rs: rs.id), start=1):
                template_name = result_set.template_id.name if result_set.template_id else 'بدون قالب'
                if template_name not in result_sets_grouped:
                    result_sets_grouped[template_name] = []
                result_sets_grouped[template_name].append(result_set)
            
            for template_name, result_sets in result_sets_grouped.items():
                worksheet.merge_cells(f'A{start_row}:H{start_row}')
                cell = worksheet[f'A{start_row}']
                cell.value = f"قالب الفحص: {template_name}"
                cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='8064A2', end_color='8064A2', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                start_row += 1
                
                for idx, result_set in enumerate(result_sets, start=1):
                    start_row = self._add_result_set_details(worksheet, start_row, result_set, idx)
                    start_row += 1
        else:
            for result_set in self.sample_id.result_set_ids:
                start_row = self._add_result_set_details(worksheet, start_row, result_set)
                start_row += 1
        
        return start_row

    # ---------------------- طابوق: تقرير مدمج ----------------------
    def _is_brick_sample(self):
        """تحديد ما إذا كانت العينة طابوق بناءً على قوالب النتائج."""
        if not self.sample_id or not self.sample_id.result_set_ids:
            return False
        for rs in self.sample_id.result_set_ids:
            code = (rs.template_id.code or '').upper() if rs.template_id else ''
            if code.startswith('BRICK_'):
                return True
        return False

    def _is_concrete_sample(self):
        if not self.sample_id or not self.sample_id.result_set_ids:
            return False
        for rs in self.sample_id.result_set_ids:
            if getattr(rs, 'is_concrete_sample', False):
                return True
            code = (rs.template_id.code or '').upper() if rs.template_id else ''
            if code.startswith('CONCRETE_'):
                return True
        return False

    def _add_brick_consolidated_report(self, ws, start_row):

        subtype_suffix = (self.sample_id.sample_subtype_id.code or '').strip().upper() if self.sample_id.sample_subtype_id else ''
        if subtype_suffix not in ('A', 'B', 'C'):
            subtype_suffix = 'A'

        ws.merge_cells(f'A{start_row}:E{start_row}')
        h = ws[f'A{start_row}']
        h.value = 'الأبعاد (متوسطات)'
        h.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        h.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        h.alignment = Alignment(horizontal='center', vertical='center')
        start_row += 1

        dim_headers = ['الطول (±3%)mm', 'العرض (±3%)mm', 'السمك (±4%)mm', 'نسبة الثقوب %', 'نسبة التثلم %']
        for col, header in enumerate(dim_headers, 1):
            c = ws.cell(row=start_row, column=col)
            c.value = header
            c.font = Font(name='Arial', size=11, bold=True, color='000000')
            c.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        start_row += 1

        avg_len = self._get_sample_first_value([f'AVG_LENGTH_{subtype_suffix}'])
        avg_wid = self._get_sample_first_value([f'AVG_WIDTH_{subtype_suffix}'])
        avg_hgt = self._get_sample_first_value([f'AVG_HEIGHT_{subtype_suffix}'])
        void_ratio = self._get_sample_first_value([f'VOID_RATIO_{subtype_suffix}'])
        chipping = self._get_sample_first_value([f'CHIPPING_PCT_{subtype_suffix}'])

        dim_values = [
            self._fmt(avg_len), self._fmt(avg_wid), self._fmt(avg_hgt), self._fmt(void_ratio, 1), self._fmt(chipping, 1)
        ]
        for col, value in enumerate(dim_values, 1):
            c = ws.cell(row=start_row, column=col)
            c.value = value
            c.font = Font(name='Arial', size=10)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        start_row += 2

        ws.merge_cells(f'A{start_row}:E{start_row}')
        h2 = ws[f'A{start_row}']
        h2.value = 'المظهر العام'
        h2.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        h2.fill = PatternFill(start_color='4BACC6', end_color='4BACC6', fill_type='solid')
        h2.alignment = Alignment(horizontal='center', vertical='center')
        start_row += 1

        appear_headers = ['التجانس', 'الزوايا', 'الشكل']
        for col, header in enumerate(appear_headers, 1):
            c = ws.cell(row=start_row, column=col)
            c.value = header
            c.font = Font(name='Arial', size=11, bold=True)
            c.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        start_row += 1

        uni = self._get_sample_first_value([f'UNIFORMITY_{subtype_suffix}'])
        cor = self._get_sample_first_value([f'CORNERS_{subtype_suffix}'])
        shp = self._get_sample_first_value([f'SHAPE_{subtype_suffix}'])
        map_txt = lambda v, good='مطابق': ('متجانس' if good == 'متجانس' else 'مطابق') if (v == 1 or v == 1.0) else ('غير متجانس' if good == 'متجانس' else 'غير مطابق')
        appear_values = [map_txt(uni, 'متجانس'), map_txt(cor), map_txt(shp)]
        for col, value in enumerate(appear_values, 1):
            c = ws.cell(row=start_row, column=col)
            c.value = value
            c.font = Font(name='Arial', size=10)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        start_row += 2

        ws.merge_cells(f'A{start_row}:E{start_row}')
        h3 = ws[f'A{start_row}']
        h3.value = 'الضغط والامتصاص'
        h3.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        h3.fill = PatternFill(start_color='9BBB59', end_color='9BBB59', fill_type='solid')
        h3.alignment = Alignment(horizontal='center', vertical='center')
        start_row += 1

        heads = ['تحمل الضغط لطابوقة واحدة N/mm²', 'معدل تحمل الضغط N/mm²', 'امتصاص الماء لطابوقة واحدة خلال %24 ساعة', 'معدل امتصاص الماء لعشرة طابوقات خلال %24 ساعة']
        for col, header in enumerate(heads, 1):
            c = ws.cell(row=start_row, column=col)
            c.value = header
            c.font = Font(name='Arial', size=11, bold=True)
            c.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        start_row += 1

        comp_list = self._get_sample_values_list([f'INDIVIDUAL_COMP_STRENGTH_{subtype_suffix}'])
        comp_avg = self._get_sample_first_value([f'COMP_STRENGTH_{subtype_suffix}'])
        abs_list = self._get_sample_values_list([f'INDIVIDUAL_ABSORB_PCT_{subtype_suffix}'])
        abs_avg = self._get_sample_first_value([f'ABSORB_PCT_{subtype_suffix}'])

        max_rows = max(len(comp_list), len(abs_list), 1)
        for i in range(max_rows):
            v1 = self._fmt(comp_list[i]) if i < len(comp_list) else ''
            v2 = self._fmt(comp_avg) if i == 0 else ''
            v3 = self._fmt(abs_list[i], 1) if i < len(abs_list) else ''
            v4 = self._fmt(abs_avg, 1) if i == 0 else ''
            row_vals = [v1, v2, v3, v4]
            for col, value in enumerate(row_vals, 1):
                c = ws.cell(row=start_row, column=col)
                c.value = value
                c.font = Font(name='Arial', size=10)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            start_row += 1

        eff_text, eff_spec = self._get_efflorescence_text_and_spec(subtype_suffix)
        ws[f'A{start_row}'] = 'تزهر'
        ws[f'B{start_row}'] = eff_text
        ws[f'C{start_row}'] = 'حدود المواصفة'
        ws[f'D{start_row}'] = eff_spec
        for col in (1,2,3,4):
            c = ws.cell(row=start_row, column=col)
            c.font = Font(name='Arial', size=10, bold=True if col in (1,3) else False)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        start_row += 2

        comp_ind_min, comp_ind_max = self._get_spec_limits([f'INDIVIDUAL_COMP_STRENGTH_{subtype_suffix}'])
        comp_avg_min, comp_avg_max = self._get_spec_limits([f'COMP_STRENGTH_{subtype_suffix}'])
        abs_min, abs_max = self._get_spec_limits([f'ABSORB_PCT_{subtype_suffix}'])

        ws[f'A{start_row}'] = 'حدود المواصفة'
        ws[f'A{start_row}'].font = Font(name='Arial', size=11, bold=True)
        start_row += 1

        spec_rows = [
            ('حد أدنى تحمل الضغط لطابوقة واحدة (N/mm²)', self._spec_text(comp_ind_min, comp_ind_max, kind='min')),
            ('حد أدنى معدل تحمل الضغط (N/mm²)', self._spec_text(comp_avg_min, comp_avg_max, kind='min')),
            ('حد أقصى معدل امتصاص الماء (%)', self._spec_text(abs_min, abs_max, kind='max')),
            ('حد أقصى نسبة التثلم (%)', '10'),
        ]
        for label, val in spec_rows:
            ws[f'A{start_row}'] = label
            ws[f'B{start_row}'] = val
            ws[f'A{start_row}'].font = Font(name='Arial', size=10)
            ws[f'B{start_row}'].font = Font(name='Arial', size=10)
            start_row += 1

        start_row += 1
        return start_row

    # ---------------------- أدوات مساعدة للطابوق (محسّنة) ----------------------
    def _fmt(self, v, prec=2):
        """تنسيق الأرقام - استخدام الأداة المشتركة"""
        return self.env['lab.data.utils'].format_value(v, prec)

    def _get_sample_values_list(self, codes):
        """استخراج قائمة القيم - استخدام الأداة المشتركة"""
        return self.env['lab.data.utils'].get_sample_values_list(
            self.sample_id, codes, sort_by_sample=True
        )

    def _get_sample_first_value(self, codes):
        """استخراج أول قيمة - استخدام الأداة المشتركة"""
        return self.env['lab.data.utils'].get_sample_first_value(
            self.sample_id, codes
        )

    def _get_spec_limits(self, codes):
        """استخراج حدود المعيار - استخدام الأداة المشتركة"""
        return self.env['lab.data.utils'].get_criterion_limits(
            self.sample_id, codes
        )

    def _spec_text(self, min_v, max_v, kind='min'):
        if kind == 'min' and min_v is not None:
            return self._fmt(min_v)
        if kind == 'max' and max_v is not None:
            return self._fmt(max_v)
        return ''

    def _get_efflorescence_text_and_spec(self, subtype_suffix):
        """إرجاع (نص التزهر، حدود المواصفة) للطابوق.
        يقرأ قيمة EFFLOR_GRADE_* ويحوّلها إلى نص باستخدام النوع الفرعي.
        """
        values = self._get_sample_values_list([f'EFFLOR_GRADE_{subtype_suffix}'])
        val = values[0] if values else None

        subtype = self.sample_id.sample_subtype_id
        if subtype and hasattr(subtype, 'get_efflorescence_display_name'):
            eff_text = subtype.get_efflorescence_display_name(int(val)) if val is not None else ''
        else:
            mapping = {1: 'لا يوجد', 2: 'خفيف', 3: 'متوسط', 4: 'عالي'}
            eff_text = mapping.get(int(val), '') if val is not None else ''

        spec_text = ''
        allowed_list = []
        if subtype and hasattr(subtype, 'get_allowed_efflorescence_values_list'):
            try:
                allowed_list = subtype.get_allowed_efflorescence_values_list()
            except Exception:
                allowed_list = []
        if allowed_list:
            names = []
            for v in sorted(set(allowed_list)):
                names.append(subtype.get_efflorescence_display_name(v) if subtype else str(v))
            spec_text = ' أو '.join(names)
        return eff_text, spec_text

    # ---------------------- خرسانة: تقرير مقاومة الضغط ----------------------
    def _add_concrete_compression_report(self, ws, start_row):
        """يبني تقرير ضغط الخرسانة مع جدول الأبعاد Area/Volume وجداول المجموعات والوسطيات."""
        task = getattr(self.sample_id, 'task_id', False)
        is_bridges = False
        try:
            st = getattr(self.sample_id, 'concrete_sample_type_id', False)
            is_bridges = bool(st and getattr(st, 'code', '') == 'CONCRETE_BRIDGES')
        except Exception:
            is_bridges = False
        try:
            ref_general = float(task.reference_general_limit) if task and getattr(task, 'reference_general_limit', False) else None
        except Exception:
            ref_general = None
        try:
            ref_min = float(task.reference_min_limit) if task and getattr(task, 'reference_min_limit', False) else None
        except Exception:
            ref_min = None
        
        ws.merge_cells(f'A{start_row}:H{start_row}')
        head = ws[f'A{start_row}']
        head.value = 'مقاومة الضغط والكثافة لمكعبات الخرسانة'
        head.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        head.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        head.alignment = Alignment(horizontal='center', vertical='center')
        start_row += 1

        dim_labels = ['الطول (مم)', 'العرض (مم)', 'الارتفاع (مم)', 'المساحة (مم²)', 'الحجم (م³)']
        dim_values = ['150', '150', '150', '22500', '0.0034']
        for idx, label in enumerate(dim_labels):
            ws[f'A{start_row}'] = label
            ws[f'A{start_row}'].font = Font(name='Arial', size=10, bold=True)
            for c in range(1, 7):
                cell = ws.cell(row=start_row, column=c+1)
                cell.value = dim_values[idx]
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            start_row += 1
        start_row += 1

        headers = ['الرمز الحقلي', 'الوزن (كغ)', 'الكثافة (كغ/م³)', 'أقصى حمل (kN)', 'مقاومة الضغط (N/mm²)', 'مقاومة الضغط المصححة (N/mm²)']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = h
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        start_row += 1

        result_sets = self.sample_id.result_set_ids.filtered(
            lambda rs: getattr(rs, 'is_concrete_sample', False) and getattr(rs, 'concrete_age_days', False) == '28'
        ).sorted('concrete_group_no')
        overall_comp_values = []
        group_averages = []

        excel_failed = False
        for rs in result_sets:
            
            weights = self._get_values_for_set(rs, 'WEIGHT_CONCRETE')
            densities = self._get_values_for_set(rs, 'DENSITY_CONCRETE')
            loads = self._get_values_for_set(rs, 'LOAD_KN_CONCRETE')
            strengths = self._get_values_for_set(rs, 'COMP_STRENGTH_CONCRETE')
            corrected = [ (v or 0) * (rs.cube_factor or 1.0) for v in strengths ]

            max_len = max(len(weights), len(densities), len(loads), len(strengths), 1)
            for i in range(max_len):
                row_vals = [
                    rs.concrete_field_serial or rs.concrete_field_code or rs.name,
                    self._fmt(weights[i] if i < len(weights) else None, 3),
                    self._fmt(densities[i] if i < len(densities) else None, 0),
                    self._fmt(loads[i] if i < len(loads) else None, 1),
                    self._fmt(strengths[i] if i < len(strengths) else None, 2),
                    self._fmt(corrected[i] if i < len(corrected) else None, 2),
                ]
                
                if i < len(strengths) and strengths[i] is not None:
                    try:
                        overall_comp_values.append(float(strengths[i]))
                    except Exception:
                        pass

                for col, v in enumerate(row_vals, 1):
                    cell = ws.cell(row=start_row, column=col)
                    cell.value = v
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(name='Arial', size=10)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                start_row += 1

            
            grp_avg_line = rs.result_line_ids.filtered(lambda l: l.criterion_id.code == 'AVG_COMP_STRENGTH_CONCRETE')
            grp_avg = grp_avg_line[0].value_numeric if grp_avg_line else (sum(strengths)/len(strengths) if strengths else 0)
            group_averages.append((rs.concrete_group_no or 0, grp_avg))
            ws[f'A{start_row}'] = f"معدل المجموعة {rs.concrete_group_no or ''}"
            ws[f'A{start_row}'].font = Font(name='Arial', size=10, bold=True)
            ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            val_disp = self._fmt(grp_avg, 2)
            try:
                if (ref_min is not None) and (grp_avg < (ref_min - min_margin)):
                    val_disp = f"{val_disp}*"
                    excel_failed = True
            except Exception:
                pass
            ws[f'E{start_row}'] = val_disp
            ws[f'E{start_row}'].font = Font(name='Arial', size=10, bold=True)
            ws[f'E{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            if '*' in val_disp:
                ws[f'E{start_row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            for c in range(1, 7):
                cell = ws.cell(row=start_row, column=c)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            start_row += 1

        if not is_bridges:

            try:
                general_margin = float(self.env['ir.config_parameter'].sudo().get_param('appointment_products.general_limit_margin', default='3'))
            except Exception:
                general_margin = 3.0

            ws.merge_cells(f'A{start_row}:H{start_row}')
            title_ref = ws[f'A{start_row}']
            title_ref.value = 'المراجع'
            title_ref.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            title_ref.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
            title_ref.alignment = Alignment(horizontal='center', vertical='center')
            start_row += 1

            labels = ['المعيار العام', 'المعيار الأدنى', 'العتبة (المعيار العام + الهامش)']
            values = [
                ('{0:.2f}'.format(ref_general)).rstrip('0').rstrip('.') if ref_general is not None else '',
                ('{0:.2f}'.format(ref_min)).rstrip('0').rstrip('.') if ref_min is not None else '',
                ('{0:.2f}'.format((ref_general or 0) + general_margin)).rstrip('0').rstrip('.') if ref_general is not None else ''
            ]
            ws[f'A{start_row}'] = labels[0]
            ws[f'B{start_row}'] = values[0]
            ws[f'C{start_row}'] = labels[1]
            ws[f'D{start_row}'] = values[1]
            ws[f'E{start_row}'] = labels[2]
            ws[f'F{start_row}'] = values[2]
            for c in ('A','C','E'):
                cell = ws[f'{c}{start_row}']
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for c in ('B','D','F'):
                cell = ws[f'{c}{start_row}']
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            start_row += 1

            start_row = self._add_group_rolling_averages(ws, start_row, reference_general_limit=ref_general)

        u_line = None
        for rs in self.sample_id.result_set_ids.filtered(
            lambda r: getattr(r, 'is_concrete_sample', False) and getattr(r, 'concrete_age_days', False) == '28'
        ):
            u_line = rs.result_line_ids.filtered(lambda l: l.criterion_id.code == 'UNCERTAINTY_STRESS_VALUE')
            if u_line:
                u_line = u_line[0]
                break
        if u_line and u_line.value_numeric is not None:
            ws[f'A{start_row}'] = 'Uncertainty Stress Value'
            ws[f'E{start_row}'] = self._fmt(u_line.value_numeric, 3)
            ws[f'A{start_row}'].font = Font(name='Arial', size=10, bold=True)
            ws[f'E{start_row}'].font = Font(name='Arial', size=10, bold=True)
            start_row += 1

        final_failed = excel_failed or bool(getattr(self, '_excel_failure_flag', False))
        ws[f'A{start_row}'] = 'نتيجة الفحص'
        ws[f'B{start_row}'] = 'فاشل' if final_failed else 'ناجح'
        ws[f'A{start_row}'].font = Font(name='Arial', size=11, bold=True)
        ws[f'B{start_row}'].font = Font(name='Arial', size=11, bold=True)
        ws[f'B{start_row}'].fill = PatternFill(start_color=('FFC7CE' if final_failed else 'C6EFCE'), end_color=('FFC7CE' if final_failed else 'C6EFCE'), fill_type='solid')
        start_row += 1

        start_row += 1
        return start_row

    def _get_values_for_set(self, result_set, code):
        """إرجاع قائمة قيم معيار محدد داخل مجموعة نتائج واحدة بالترتيب - استخدام الأدوات المشتركة"""
        return self.env['lab.data.utils'].get_result_set_values_list(
            result_set, code, sort_by_sample=True
        )

    def _add_result_set_details(self, worksheet, start_row, result_set, group_idx=None):
        """إضافة تفاصيل مجموعة نتائج واحدة"""
        worksheet[f'A{start_row}'] = f"مجموعة النتائج: {result_set.name}"
        worksheet[f'A{start_row}'].font = Font(name='Arial', size=12, bold=True)
        worksheet[f'D{start_row}'] = f"الحالة: {dict(result_set._fields['state'].selection).get(result_set.state, '')}"
        worksheet[f'F{start_row}'] = f"النتيجة: {dict(result_set._fields['overall_result'].selection).get(result_set.overall_result, '')}"
        
        if result_set.overall_result == 'pass':
            worksheet[f'F{start_row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif result_set.overall_result == 'fail':
            cell = worksheet[f'F{start_row}']
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        start_row += 1
        
        try:
            template_code = (result_set.template_id.code or '').upper() if result_set.template_id else ''
            industry_type = (result_set.template_id.industry_type or '').lower() if result_set.template_id else ''
        except Exception:
            template_code = ''
            industry_type = ''

        if template_code == 'COARSE_SIEVE' or industry_type == 'aggregate':
            start_row = self._add_coarse_aggregate_sieve_table(worksheet, start_row, result_set)
            return start_row

        if result_set.is_concrete_sample:
            display_group_no = group_idx if group_idx is not None else result_set.concrete_group_no
            concrete_info = [
                ('رقم مجموعة الخرسانة', display_group_no or ''),
                ('عمر الخرسانة', dict(result_set._fields['concrete_age_days'].selection).get(result_set.concrete_age_days, '')),
                ('الرمز المختبري', result_set.concrete_field_code or ''),
                ('الرمز الحقلي', result_set.concrete_field_serial or ''),
                ('رقم الدفعة', result_set.batch_number or ''),
                ('تاريخ الصب', result_set.casting_date.strftime('%Y-%m-%d') if result_set.casting_date else ''),
                ('تاريخ الفحص', result_set.testing_date.strftime('%Y-%m-%d') if result_set.testing_date else ''),
                ('العمر المطلوب للفحص', dict(result_set._fields['required_age_days'].selection).get(result_set.required_age_days, '')),
                ('درجة حرارة الفحص (°C)', ('{0:.2f}'.format(result_set.testing_temperature)).rstrip('0').rstrip('.') if result_set.testing_temperature else ''),
                ('معامل المكعب', ('{0:.2f}'.format(result_set.cube_factor)).rstrip('0').rstrip('.') if result_set.cube_factor else '')
            ]
            for label, value in concrete_info:
                worksheet[f'A{start_row}'] = label
                worksheet[f'B{start_row}'] = value
                worksheet[f'A{start_row}'].font = Font(name='Arial', size=11, bold=True)
                worksheet[f'B{start_row}'].font = Font(name='Arial', size=11)
                start_row += 1
            start_row += 1
        
        headers = ['رقم العينة', 'المعيار', 'الوحدة', 'الحد الأدنى', 'الحد الأعلى', 'القيمة', 'المطابقة', 'الملاحظات']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col)
            cell.value = header
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        start_row += 1
        
        def _line_sort_key(l):
            if l.sample_identifier == 'عام':
                primary = 2  
            elif l.data_type == 'computed' and not getattr(l.criterion_id, 'is_global', False):
                primary = 1  
            else:
                primary = 0  
            return (primary, l.summary_sort_key or '', l.criterion_name or '')
        for line in sorted(result_set.result_line_ids, key=_line_sort_key):
            self._add_result_line(worksheet, start_row, line)
            start_row += 1
        
        return start_row

    # -------- منطق خاص بجدول المناخل للركام الخشن --------
    def _add_coarse_aggregate_sieve_table(self, ws, start_row, result_set):
        """إضافة جدول تحليل منخلي للركام الخشن بنفس ترتيب الصورة.
        يعتمد على الأكواد المعرفة في template COARSE_SIEVE.
        """

        ws.merge_cells(f'A{start_row}:H{start_row}')
        cell = ws[f'A{start_row}']
        cell.value = 'تحليل منخلي للركام الخشن'
        cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='8064A2', end_color='8064A2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        start_row += 1

        orig_wt = self._get_code_value_numeric(result_set, 'ORIG_WT_GM')
        ws[f'A{start_row}'] = 'وزن العينة الأصلية (غم)'
        ws[f'B{start_row}'] = ('{0:.0f}'.format(orig_wt)).rstrip('0').rstrip('.') if orig_wt else ''
        ws[f'A{start_row}'].font = Font(name='Arial', size=11, bold=True)
        ws[f'B{start_row}'].font = Font(name='Arial', size=11)
        start_row += 1

        headers = ['Sieve (mm)', 'Sieve (in)', 'Wt Retained (gm)', 'Retained %', 'Passing %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        start_row += 1

        sieves = [
            {'mm': '75', 'in': '3', 'cum_pct': 'CUM_RET_75MM_PCT'},
            {'mm': '63', 'in': '2½', 'cum_pct': 'CUM_RET_63MM_PCT'},
            {'mm': '37.5', 'in': '1½', 'cum_pct': 'CUM_RET_37_5MM_PCT'},
            {'mm': '20', 'in': '', 'cum_pct': 'CUM_RET_20MM_PCT'},
            {'mm': '14', 'in': '', 'cum_pct': 'CUM_RET_14MM_PCT'},
            {'mm': '10', 'in': '', 'cum_pct': 'CUM_RET_10MM_PCT'},
            {'mm': '5', 'in': 'No.8', 'cum_pct': 'CUM_RET_5MM_PCT'},
            {'mm': '2.36', 'in': '', 'cum_pct': 'CUM_RET_2_36MM_PCT'},
            {'mm': '0.075', 'in': 'No.200', 'cum_pct': 'CUM_RET_0_075MM_PCT'},
        ]

        for sieve in sieves:
            cum_ret_pct = self._get_code_value_numeric(result_set, sieve['cum_pct'])
            cum_ret_gm = (orig_wt * cum_ret_pct / 100.0) if orig_wt else 0.0
            passing_pct = 100.0 - cum_ret_pct if cum_ret_pct is not None else 0.0

            row_values = [
                sieve['mm'],
                sieve['in'],
                ('{0:.1f}'.format(cum_ret_gm)).rstrip('0').rstrip('.') if cum_ret_gm or cum_ret_gm == 0 else '',
                ('{0:.2f}'.format(cum_ret_pct)).rstrip('0').rstrip('.') if cum_ret_pct or cum_ret_pct == 0 else '',
                ('{0:.2f}'.format(passing_pct)).rstrip('0').rstrip('.') if passing_pct or passing_pct == 0 else '',
            ]

            for col, value in enumerate(row_values, 1):
                cell = ws.cell(row=start_row, column=col)
                cell.value = value
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            start_row += 1

        start_row += 1

        ws.merge_cells(f'A{start_row}:E{start_row}')
        title_cell = ws[f'A{start_row}']
        title_cell.value = 'المواد الخشنة و الناعمة'
        title_cell.font = Font(name='Arial', size=11, bold=True, color='000000')
        title_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        start_row += 1

        fines_headers = ['نوع الفحص', 'نتيجة الفحص', 'حدود المواصفة']
        for col, header in enumerate(fines_headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = Font(name='Arial', size=11, bold=True, color='000000')
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        start_row += 1

        fines_pct = self._get_code_value_numeric(result_set, 'FINES_0_075MM_WASH_PCT')
        if (fines_pct is None or fines_pct == 0) and orig_wt:
            washed_wt = self._get_code_value_numeric(result_set, 'WASHED_WT_GM')
            fines_pct = ((orig_wt - washed_wt) / orig_wt * 100.0) if orig_wt else 0.0

        fines_label = 'المواد المارة من منخل % (0.075)mm'
        fines_value_disp = ('{0:.2f}'.format(fines_pct)).rstrip('0').rstrip('.') if fines_pct is not None else ''
        spec_limit = 'حد أعلى 3'

        fines_row = [fines_label, fines_value_disp, spec_limit]
        for col, value in enumerate(fines_row, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = value
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        start_row += 2
        return start_row

    def _get_code_value_numeric(self, result_set, code):
        """إرجاع القيمة الرقمية لسطر بمعيار code.
        يدعم الحقول المحسوبة وغير المحسوبة.
        """
        try:
            line = result_set.result_line_ids.filtered(lambda l: l.criterion_id and l.criterion_id.code == code)[:1]
            if not line:
                return 0.0
            line = line[0]
            val = None
            try:
                val = line.result_value_computed
            except Exception:
                val = None
            if val in (None, False, 0) and hasattr(line, 'value_numeric'):
                val = line.value_numeric or 0.0
            return float(val or 0.0)
        except Exception:
            return 0.0

    def _add_result_line(self, worksheet, row, result_line):
        """إضافة سطر نتيجة واحد"""

        if result_line.data_type == 'numeric':
            if result_line.result_value_numeric is not None:
                display_value = ('{0:.3f}'.format(result_line.result_value_numeric)).rstrip('0').rstrip('.')
            else:
                display_value = ''
        elif result_line.data_type == 'text':
            display_value = result_line.result_value_text or ''
        elif result_line.data_type == 'selection':
            display_value = result_line.result_value_selection or ''
        elif result_line.data_type == 'boolean':
            display_value = 'نعم' if result_line.result_value_boolean else 'لا'
        elif result_line.data_type == 'computed':
            if result_line.result_value_computed is not None:
                display_value = ('{0:.3f}'.format(result_line.result_value_computed)).rstrip('0').rstrip('.')
            else:
                display_value = ''
        elif result_line.data_type == 'date':
            display_value = result_line.result_value_date.strftime('%Y-%m-%d') if result_line.result_value_date else ''
        else:
            display_value = ''
        
        is_summary = getattr(result_line.criterion_id, 'is_summary_field', False)
        is_computed_non_global = result_line.data_type == 'computed' and not is_summary and not getattr(result_line.criterion_id, 'is_global', False)
        if result_line.sample_identifier == 'عام':
            sample_id_display = 'النتيجة'
        elif is_computed_non_global:
            sample_id_display = f"النتيجة {result_line.sample_identifier or ''}"
        else:
            sample_id_display = result_line.sample_identifier or ''

        conformity_text = {
            'pass': 'مطابق ✓',
            'fail': 'غير مطابق ✗',
            'pending': 'قيد الانتظار ⏳'
        }.get(result_line.conformity_status, '')
        
        data = [
            sample_id_display,
            result_line.criterion_name or '',
            result_line.unit_of_measure or '',
            result_line.min_value if result_line.min_value is not None else '',
            result_line.max_value if result_line.max_value is not None else '',
            display_value,
            conformity_text,
            result_line.notes or ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = value
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if col == 1 and (result_line.sample_identifier == 'عام' or is_computed_non_global):
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            if col == 7:
                if result_line.conformity_status == 'pass':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif result_line.conformity_status == 'fail':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                elif result_line.conformity_status == 'pending':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    def _add_group_rolling_averages(self, worksheet, start_row, reference_general_limit=None, window_size=4):
        """Add table of consecutive concrete groups averages with variable window size (default 4).
        - If reference_general_limit is provided: Put asterisk * when the window average is less than (general limit + 3).
        - The title shows the window size dynamically.
        """
        try:
            sample_type = getattr(self.sample_id, 'concrete_sample_type_id', False)
            if sample_type and getattr(sample_type, 'code', '') == 'CONCRETE_BRIDGES':
                return start_row
        except Exception:
            pass
        TARGET_CRIT = 'متوسط قوة الضغط (نيوتن/مم²)'
        values = []
        result_sets = self.sample_id.result_set_ids.filtered(
            lambda rs: getattr(rs, 'is_concrete_sample', False) and getattr(rs, 'concrete_age_days', False) == '28'
        ).sorted('concrete_group_no')
        for rs in result_sets:
            line = rs.result_line_ids.filtered(lambda l: l.criterion_name == TARGET_CRIT)
            if line:
                val = line[0].result_value_computed or line[0].result_value_numeric
                if val is not None:
                    values.append(val)
        if len(values) < window_size:
            return start_row
        

        worksheet.merge_cells(f'A{start_row}:H{start_row}')
        title_cell = worksheet[f'A{start_row}']
        title_cell.value = 'متوسطات مجموعات الخرسانة المتتالية'
        title_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        start_row += 1

        worksheet[f'A{start_row}'] = 'المجموعات'
        worksheet[f'B{start_row}'] = 'متوسط القوة (N/mm²)'
        for col in (1,2):
            cell = worksheet.cell(row=start_row, column=col)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        start_row += 1
        
        window = window_size
        # قراءة هامش المعيار العام من الإعدادات
        try:
            general_margin = float(self.env['ir.config_parameter'].sudo().get_param('appointment_products.general_limit_margin', default='3'))
        except Exception:
            general_margin = 3.0

        for idx in range(0, len(values)-window+1):
            rng = f"{idx+1}-{idx+window}"
            avg_val = sum(values[idx:idx+window])/window
            worksheet[f'A{start_row}'] = rng
            disp = ('{0:.2f}'.format(avg_val)).rstrip('0').rstrip('.')
            try:

                if (reference_general_limit is not None) and (avg_val < (float(reference_general_limit) + general_margin)):
                    disp = f"{disp}*"
                    self._excel_failure_flag = True
            except Exception:
                pass
            worksheet[f'B{start_row}'] = disp
            for col in (1,2):
                cell = worksheet.cell(row=start_row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            start_row += 1
        
        start_row += 1
        return start_row



    def _to_plain_text(self, html):
        """تحويل HTML إلى نص بسيط لعرضه في Excel."""
        if not html:
            return ''
        try:
            from odoo.tools import html2plaintext
            return (html2plaintext(html) or '').strip()
        except Exception:
            text = re.sub(r'<[^>]+>', ' ', str(html))
            text = re.sub(r'\s+', ' ', text)
            return text.replace('&nbsp;', ' ').strip()

    def _adjust_column_widths(self, worksheet):

        try:
            max_col = worksheet.max_column or 1
            for col_idx in range(1, max_col + 1):
                column_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[column_letter].width = 9.9
        except Exception:
            pass

    def _wrap_all_cells(self, worksheet):
        try:
            max_row = worksheet.max_row or 0
            max_col = worksheet.max_column or 0
            thin_side = Side(style='thin')
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border
        except Exception:
            pass

    def _apply_outer_thick_borders(self, worksheet):
        
        try:
            max_row = worksheet.max_row or 0
            max_col = worksheet.max_column or 0
            thick = Side(style='thick')
            thin = Side(style='thin')
            for top in range(1, max_row + 1):

                prev_empty = True if top == 1 else all((worksheet.cell(row=top-1, column=c).value in (None, '')) for c in range(1, max_col+1))
                row_empty = all((worksheet.cell(row=top, column=c).value in (None, '')) for c in range(1, max_col+1))
                if prev_empty and not row_empty:

                    bottom = top
                    while bottom <= max_row and not all((worksheet.cell(row=bottom, column=c).value in (None, '')) for c in range(1, max_col+1)):
                        bottom += 1
                    bottom -= 1
                    if bottom < top:
                        continue
                    for r in range(top, bottom+1):
                        for c in range(1, max_col+1):
                            cell = worksheet.cell(row=r, column=c)
                            left = thick if c == 1 else thin
                            right = thick if c == max_col else thin
                            top_side = thick if r == top else thin
                            bottom_side = thick if r == bottom else thin
                            cell.border = Border(left=left, right=right, top=top_side, bottom=bottom_side)

                    top = bottom + 1
        except Exception:
            pass

    def _merge_empty_with_neighbors(self, worksheet):

        try:
            max_row = worksheet.max_row or 0
            max_col = worksheet.max_column or 0
            r = 1
            while r <= max_row:
                c = 1
                while c <= max_col:
                    cell = worksheet.cell(row=r, column=c)
                    if (cell.value in (None, '')) and c < max_col:
                        k = c + 1
                        while k <= max_col and (worksheet.cell(row=r, column=k).value in (None, '')):
                            k += 1
                        if k <= max_col and (worksheet.cell(row=r, column=k).value not in (None, '')):
                            start = get_column_letter(c)
                            end = get_column_letter(k)
                            worksheet.merge_cells(f"{start}{r}:{end}{r}")
                            c = k
                        else:
                            c += 1
                    else:
                        c += 1
                r += 1
        except Exception:
            pass

    def _apply_zebra_striping(self, worksheet, start_row=1):

        try:
            max_row = worksheet.max_row or 0
            max_col = worksheet.max_column or 0
            alt_fill = PatternFill(start_color='F7F7F7', end_color='F7F7F7', fill_type='solid')
            for r in range(start_row, max_row + 1):
                if r % 2 == 1:
                    for c in range(1, max_col + 1):
                        worksheet.cell(row=r, column=c).fill = worksheet.cell(row=r, column=c).fill or alt_fill
        except Exception:
            pass

    def _finalize_sheet(self, worksheet):

        try:
            worksheet.sheet_view.showGridLines = False
        except Exception:
            pass
        self._wrap_all_cells(worksheet)
        self._apply_zebra_striping(worksheet, start_row=1)
        self._adjust_column_widths(worksheet)
        try:
            worksheet.sheet_format.defaultRowHeight = 18
        except Exception:
            pass

        try:
            worksheet.freeze_panes = None
        except Exception:
            pass

        try:
            worksheet.print_title_rows = '1:6'
            worksheet.header_footer.center_header = "&Bتقرير نتائج فحص العينة&B"
            worksheet.header_footer.left_footer = (self.sample_id.name or '')
            worksheet.header_footer.right_footer = "صفحة &P من &N"
            from openpyxl.utils import get_column_letter as _gcl
            worksheet.print_area = f"A1:{_gcl(worksheet.max_column)}{worksheet.max_row}"
        except Exception:
            pass
        self._apply_outer_thick_borders(worksheet)

    def action_download_file(self):
        """تحميل الملف المُصدر"""
        self.ensure_one()
        if not self.export_file:
            raise UserError(_("لا يوجد ملف للتحميل. يرجى تصدير البيانات أولاً."))
        
        download_url = f'/web/content?model={self._name}&id={self.id}&field=export_file&download=true&filename={self.export_filename}'
        
        return {
            'type': 'ir.actions.act_url',
            'url': download_url,
            'target': 'self',
        }

    def action_reset(self):
        """إعادة تعيين المعالج"""
        self.ensure_one()
        self.write({
            'export_file': False,
            'export_filename': False,
            'state': 'draft'
        })
        return {'type': 'ir.actions.act_window_close'}
